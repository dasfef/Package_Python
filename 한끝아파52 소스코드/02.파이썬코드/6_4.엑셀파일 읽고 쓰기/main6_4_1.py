from openpyxl import load_workbook
from openpyxl import Workbook

save_file_path = r"6_4.엑셀파일 읽고 쓰기\엑셀에쓰기.xlsx"

name_list = ["장다인","임꺽정","홍길동"]

try:
    wb = load_workbook(save_file_path, data_only=True)
    sheet  = wb.active
except:
    wb = Workbook()
    sheet = wb.active

for name in name_list:
    sheet.append([name])

wb.save(save_file_path)