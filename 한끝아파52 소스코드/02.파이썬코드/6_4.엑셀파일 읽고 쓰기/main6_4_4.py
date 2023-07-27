from openpyxl import load_workbook
from openpyxl import Workbook

save_file_path = r"6_4.엑셀파일 읽고 쓰기\엑셀에쓰기.xlsx"

try:
    wb = load_workbook(save_file_path, data_only=True)
    sheet  = wb.active
except:
    wb = Workbook()
    sheet = wb.active

print("A2의 값:",sheet['A2'].value)

print("4행 2열의 값:",sheet.cell(4, 2).value)

print("="*30)
print("A1부터 A5까지의 값:")
cell_list = sheet['A1' : 'A5']
for row in cell_list:
    for cell in row:
        print(cell.value)
print("="*30)