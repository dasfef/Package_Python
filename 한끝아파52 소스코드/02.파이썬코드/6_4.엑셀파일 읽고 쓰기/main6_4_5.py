from openpyxl import load_workbook
from openpyxl import Workbook

save_file_path = r"6_4.엑셀파일 읽고 쓰기\엑셀에쓰기.xlsx"

try:
    wb = load_workbook(save_file_path, data_only=True)
    sheet  = wb.active
except:
    wb = Workbook()
    sheet = wb.active

name_list = []
for i in range(1,sheet.max_row + 1):
    name_list.append(sheet.cell(i, 1).value)

print(name_list)