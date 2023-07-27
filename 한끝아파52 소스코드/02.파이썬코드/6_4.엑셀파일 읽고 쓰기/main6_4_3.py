from openpyxl import load_workbook
from openpyxl import Workbook

save_file_path = r"6_4.엑셀파일 읽고 쓰기\엑셀에쓰기.xlsx"

try:
    wb = load_workbook(save_file_path, data_only=True)
    sheet  = wb.active
except:
    wb = Workbook()
    sheet = wb.active
    

sheet.append([1,2,3,"값추가"])

sheet['C6'] = '안녕하세요'
sheet['D6'] = 123

sheet.cell(2, 7, '2행7열')

wb.save(save_file_path)